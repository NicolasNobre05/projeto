import openpyxl
import os

import openpyxl.worksheet
from menus import Menu
from datetime import datetime
from pessoas import Pessoa
from produto import Produto
import random

#DECLARANDO VARIAVEL PLANILHA VENDAS
vendasCadastrados = openpyxl.load_workbook('Vendas.xlsx')
vendasCadastradoPlanilha1 = vendasCadastrados['Planilha1']
vendasCadastradoSheet1 = vendasCadastrados['Sheet1']
dfVendasCadastrados = 'Vendas.xlsx'

#DECLARANDO VARIAVEL PLANILHA PRODUTO
produtosCadastrados = openpyxl.load_workbook('Produtos.xlsx')
dfProdutosCadastrados = 'Produtos.xlsx'

#DECLARANDO VARIAVEL PLANILHA CLIENTE
clientesCadastrados = openpyxl.load_workbook('Clientes.xlsx')
dfClientesCadastrados = 'Clientes.xlsx'

#DECLARANDO VARIAVEL PLANILHA VENDEDORES
VendedoresCadastrados = openpyxl.load_workbook('Vendedores.xlsx')
dfVendedoresCadastrados = 'Vendedores.xlsx'

#ATIVANDO AS PLANILHAS
vendasCadastrado = vendasCadastrados.active 

produtosCadastrado = produtosCadastrados.active  

clienteCadastrado = clientesCadastrados.active

vendedorCadastrado = VendedoresCadastrados.active



class Venda:

    #FUNÇÕES PRINCIPAIS

    def __init__(self,id_Vendas, vendedor,vendedorResp, cliente, clienteResp, tipoPagamento, dataCompra,quantidadeTotalProdutos, valorTotalCompra, produtos, quantidades, ValoresTotalProdutos):
        self.id = id_Vendas
        self.vendedor = vendedor
        self.nomeVendedor = vendedorResp
        self.cliente = cliente
        self.nomeCliente = clienteResp
        self.tipoDePagamento = tipoPagamento
        self.dataCompra = dataCompra
        self.quantidade = quantidadeTotalProdutos
        self.valorTotal = valorTotalCompra
        self.produtos = produtos
        self.quantidadeTotalProdutos = quantidades
        self.valorTotalProduto = ValoresTotalProdutos

    def exibirVenda(self):
        print("\033[1;32mDetalhes da Venda:\033[0m")
        print(f"ID da Venda: {self.id}")
        print(f"Vendedor: {self.nomeVendedor}")
        print(f"Cliente: {self.nomeCliente}")
        print(f"Data da Compra: {self.dataCompra.strftime('%d/%m/%Y') if isinstance(self.dataCompra, datetime) else self.dataCompra}")
        print(f"Quantidade Total de Produtos: {self.quantidade}")
        print(f"Valor Total da Compra: R$ {self.valorTotal:.2f}")

    def menuVendas():
        os.system('cls')
        print("\033[1;34m" + "----------MENU VENDAS----------" + "\033[0m")
        print("\033[1;33m" + "[1] Nova venda" + "\033[0m")
        print("\033[1;33m" + "[2] Histórico de vendas" + "\033[0m")
        print("\033[1;33m" + "[3] Voltar ao menu principal" + "\033[0m")
        print("\033[1;34m" + "---------------------------------" + "\033[0m")
        
        escolhaMenuVendas = int(input("Selecione uma opção: "))
        
        if escolhaMenuVendas == 1:
            os.system('cls')
            Venda.opcaoNovaVenda()
        elif escolhaMenuVendas == 2:
            os.system('cls')
            Venda.opcaoHistoricoVenda()
        elif escolhaMenuVendas == 3:
            os.system('cls')
            Menu.menuPrincipal()
        else:
            from common import main
            os.system('cls')
            print("Opção invalida...")
            main()

    def opcaoNovaVenda():
        from common import main

        print("\033[1;34m" + "-----------------------------------NOVA VENDA-----------------------------------" + "\033[0m")
        
        vendedor, vendedorResp = Venda.vendedorDaVenda()
        
        cliente, clienteResp = Venda.clienteDaVenda(clienteCadastrado)

        dataCompra = datetime.now()

        produtos,nomeProdutosComprados, quantidades, ValoresTotalProdutos, quantidadeTotalProdutos, valorTotalCompra = Venda.produtosComprados(produtosCadastrado, produtosCadastrados)
        
        print("\033[1;34m" + "-----------------------------------COMPRA FINALIZADA-----------------------------------" + "\033[0m")
        print("\033[1;32mVendedor:  \033[0m", vendedorResp)
        print("\033[1;32mCliente:  \033[0m", clienteResp)
        print("\033[1;32mData da compra:  \033[0m", dataCompra.strftime('%d/%m/%Y'))
        
        for produto, quantidade, preco in zip(nomeProdutosComprados, quantidades, ValoresTotalProdutos):
            print(f"\033[1;32mProduto:\033[0m {produto} | \033[1;32mQuantidade:\033[0m {quantidade} | \033[1;32mPreço Total:\033[0m R${preco}")
        print("\033[1;32mTotal itens:  \033[0m", quantidadeTotalProdutos)
        print("\033[1;32mTotal compra:  \033[0m", valorTotalCompra)
        
        print("\033[1;33m" + "[1] Dinheiro" + "\033[0m")
        print("\033[1;33m" + "[2] PIX" + "\033[0m")
        print("\033[1;33m" + "[3] Cartão de debito" + "\033[0m")
        print("\033[1;33m" + "[4] Cartão de credito" + "\033[0m")
        
        tipoPagamento = Venda.tipoPagamento()
        ID_Vendas = Venda.geradorCadastro()
        venda = Venda(ID_Vendas, vendedor,vendedorResp, cliente,clienteResp, tipoPagamento, dataCompra, quantidadeTotalProdutos, valorTotalCompra, produtos, quantidades, ValoresTotalProdutos)
        
        Venda_Dados = {
            'ID': venda.id,
            'Vendedor': venda.vendedor,
            'Cliente': venda.cliente,
            'Tipo_Pagamento': venda.tipoDePagamento,
            'Data_Pagamento': venda.dataCompra,
            'Quantidade': venda.quantidade,
            'Valor_Total': venda.valorTotal
        }

        Venda_DadosProdutos = {
            'ID': venda.id,
            'Produto': venda.produtos,
            'Quantidade': venda.quantidadeTotalProdutos,
            'Preço total': venda.valorTotalProduto
        }

        vendasCadastradoSheet1.append(list(Venda_Dados.values()))

        for i in range(len(Venda_DadosProdutos['Produto'])):
            linha = [
                Venda_DadosProdutos['ID'],                  
                Venda_DadosProdutos['Produto'][i],          
                Venda_DadosProdutos['Quantidade'][i],       
                Venda_DadosProdutos['Preço total'][i]       
            ]
            vendasCadastradoPlanilha1.append(linha)
        
        vendasCadastrados.save('Vendas.xlsx')
        venda.exibirVenda()

    def opcaoHistoricoVenda():
        planilhas = dfVendasCadastrados
        verifPlanilha = 1
        Venda.imprimirPlanilha(planilhas, verifPlanilha)
        Menu.menuVenda()


    #FUNÇÕES AUXILIARES

    def vendedorDaVenda():
        while True:

            planilhas = dfVendedoresCadastrados
            verifPlanilha = 3
            Venda.imprimirPlanilha(planilhas, verifPlanilha)
        
            vendedorResp = input("\033[1;32mCadastro do vendedor: \033[0m").upper()
            print(vendedorResp)
            if vendedorResp[:2] != "VR":
                os.system('cls')
                print("\033[31mCadastro invalido ....\033[0m")
            else:
                break

        celulaProcurada = vendedorResp
        planilhas = vendedorCadastrado

        Venda.procurarCelulaVendas(celulaProcurada,planilhas, verifPlanilha)
        os.system('cls')
        vendedorResp = Venda.procxVendedorClienteProduto(celulaProcurada, planilhas, verifPlanilha)
        print("\033[1;32mVendedor:  \033[0m", vendedorResp)
        Venda.somaVendasClienteVendedor(celulaProcurada, planilhas, verifSomaVenda= 3)
        return celulaProcurada, vendedorResp

    def clienteDaVenda(clienteCadastrado):

        while True:
            planilhas = dfClientesCadastrados
            verifPlanilha = 2
            Venda.imprimirPlanilha(planilhas, verifPlanilha)
            
            print("\033[1;33m" + "[2] Cliente não cadastrado" + "\033[0m") 

            clienteResp = input("\033[1;32mCadastro do cliente: \033[0m").upper()
            if clienteResp[:2] != "CL" and clienteResp != "2":
                os.system('cls')
                print("\033[31mCadastro invalido ....\033[0m")
            elif clienteResp == "2":
                os.system('cls')
                Pessoa.opcaoCadastroPessoa(opcaoMenu= 2, opcaoPessoa= "CLIENTE")
                clientesCadastrados = openpyxl.load_workbook('Clientes.xlsx')
                clienteCadastrado = clientesCadastrados.active
            else:
                break
        celulaProcurada = clienteResp
        planilhas = clienteCadastrado

        Venda.procurarCelulaVendas(celulaProcurada,planilhas, verifPlanilha)
        os.system('cls')
        clienteResp = Venda.procxVendedorClienteProduto(celulaProcurada, planilhas, verifPlanilha)
        print("\033[1;32mCliente:  \033[0m", clienteResp)
        Venda.somaVendasClienteVendedor(celulaProcurada, planilhas, verifSomaVenda= 2)
        return celulaProcurada, clienteResp

    def produtosComprados(produtosCadastrado, produtosCadastrados):
        ProdutosComprados = []
        nomeProdutosComprados = []
        QuantidadesProdutosComprados = []
        PrecoTotalProdutosComprados = []
        while True:

            while True:
                planilhas = dfProdutosCadastrados
                verifPlanilha = 4
                Venda.imprimirPlanilha(planilhas, verifPlanilha)

                print("\033[1;33m" + "[1] Finalizar compra" + "\033[0m") 
                print("\033[1;33m" + "[2] Produto não cadastrado" + "\033[0m") 

                produtoVenda = input("\033[1;32mCódigo do produto: \033[0m").upper()
                
                if produtoVenda[:2] != "AL" and produtoVenda[:2] != "HI" and produtoVenda[:2] != "OU" and produtoVenda != "2" and produtoVenda != "1" and not ProdutosComprados:
                    os.system('cls')
                    print("\033[31mCadastro invalido ....\033[0m")
                elif produtoVenda == "2":
                    os.system('cls')
                    Produto.opcaoCadastroProd()
                    produtosCadastrados = openpyxl.load_workbook('Produtos.xlsx')
                    produtosCadastrado = produtosCadastrados.active 
                else:
                    break
            
            if produtoVenda == "1":
                return ProdutosComprados, nomeProdutosComprados, QuantidadesProdutosComprados, PrecoTotalProdutosComprados,somaQuantidadeProdutos, somaPrecoTotalProdutosComprados


            celulaProcurada = produtoVenda
            planilhas = produtosCadastrado

            Venda.procurarCelulaVendas(celulaProcurada,planilhas, verifPlanilha)

            produtoVenda = Venda.procxVendedorClienteProduto(celulaProcurada, planilhas, verifPlanilha)
            print("\033[1;32mProduto:  \033[0m", produtoVenda)
            
            quantidadeProduto = int(Venda.quantidadeProdutos(celulaProcurada, planilhas, verifPlanilha, produtosCadastrados))
            
            precoTotalProduto = int(Venda.precoTotalPorProduto(celulaProcurada, quantidadeProduto, planilhas))
            
            
            ProdutosComprados.append(celulaProcurada)
            nomeProdutosComprados.append(produtoVenda)
            QuantidadesProdutosComprados.append(quantidadeProduto)
            PrecoTotalProdutosComprados.append(precoTotalProduto)
            
            somaQuantidadeProdutos = sum(QuantidadesProdutosComprados)
            somaPrecoTotalProdutosComprados = sum(PrecoTotalProdutosComprados)

    def tipoPagamento():
        tipoPagamento = str(input("Digite sua opção: "))
        if tipoPagamento == "1":
            tipoPagamento = "Dinheiro"
        elif tipoPagamento == "2":
            tipoPagamento = "PIX"
        elif tipoPagamento == "3":
            tipoPagamento = "Cartão de debito"
        elif tipoPagamento == "4":
            tipoPagamento = "Cartão de credito"
        else:
            print("Opção invalida...")
        return tipoPagamento

    def geradorCadastro():
        from common import procurarCelula
        numero = ''.join(random.choices('0123456789', k=5))
        planilhas = vendasCadastrado
        celulaProcurada = numero
        celula = procurarCelula(planilhas, celulaProcurada)

        if celula != None:
                Venda.geradorCadastro(produtosCadastrado)
                
        return numero

    def quantidadeProdutos(celulaProcurada, planilhas, verifPlanilha, produtosCadastrados):
        from common import verificarNumeros
        quantidadeProduto = input("\033[1;32mQuantidade: \033[0m")
        verifNumeros = quantidadeProduto
        verifNumeros = verificarNumeros(verifNumeros)
        if verifNumeros == False:
            print("\033[31mQuantidade invalida ....\033[0m")
        
        verifQuantidade = Venda.procxVendedorClienteProduto(celulaProcurada, planilhas, verifPlanilha)
        if verifQuantidade < quantidadeProduto:
            print("\033[31mQuantidade invalida ....\033[0m")
            Venda.quantidadeProdutos(celulaProcurada, planilhas)

        Venda.procxSubtrairEstoque(celulaProcurada, quantidadeProduto, planilhas, produtosCadastrados)
        
        return quantidadeProduto

    def precoTotalPorProduto(celulaProcurada, quantidadeProduto, planilhas):
        
        precoProduto = float(Venda.procxVendedorClienteProduto(celulaProcurada, planilhas, verifPlanilha= 5))
        TotalPrecoProduto = precoProduto * quantidadeProduto

        return TotalPrecoProduto

    def procxSubtrairEstoque(celulaProcurada, quantidadeProduto, planilhas, produtosCadastrados):
            for row in planilhas.iter_rows(min_row=1, max_row=planilhas.max_row):
                if row[4].value == celulaProcurada: 
                    Estoque =  int(row[3].value )
                    Estoque -= int(quantidadeProduto)
                    row[3].value = Estoque
                    produtosCadastrados.save('Produtos.xlsx')
                    produtosCadastrados = openpyxl.load_workbook('Produtos.xlsx')

    def somaVendasClienteVendedor(celulaProcurada,planilhas, verifSomaVenda,):
        for row in planilhas.iter_rows(min_row=1, max_row=planilhas.max_row):
                if row[3].value == celulaProcurada: 
                    quantidadeVenda =  int(row[5].value )
                    quantidadeVenda += 1
                    row[5].value = quantidadeVenda
                    if verifSomaVenda == 2:
                        clientesCadastrados.save('Clientes.xlsx')
                        clientesCadastrados = openpyxl.load_workbook('Clientes.xlsx')
                    else:
                        VendedoresCadastrados.save('Vendedores.xlsx')
                        VendedoresCadastrados = openpyxl.load_workbook('Vendedores.xlsx')

    def procxVendedorClienteProduto(celulaProcurada, planilhas, verifPlanilha ):
        linha = 0

        if verifPlanilha == 5:
            linha = 2

        for row in planilhas.iter_rows(min_row=1, max_row=planilhas.max_row):
            if row[4].value == celulaProcurada: 
                return row[linha].value  
        
    def imprimirPlanilha(planilhas, verifPlanilha):
        from common import imprimirTabelas
        
        imprimirTabelas(planilhas, verifPlanilha)
    
    def procurarCelulaVendas(celulaProcurada, planilhas, verifPlanilha):
        from common import procurarCelula

        celula = procurarCelula(planilhas, celulaProcurada)
        
        if celula == None:
            if verifPlanilha == 1:
                print("\033[31m Registro de venda não encontrado ....\033[0m")
            elif verifPlanilha == 2:
                os.system('cls')
                print("\033[31m Cliente não registrado  ....\033[0m")
                Venda.clienteDaVenda(clienteCadastrado)
            elif verifPlanilha == 3:
                print("\033[31m Vendedor não registrado  ....\033[0m")
                Venda.vendedorDaVenda()
            elif verifPlanilha == 4:
                print("\033[31m Produto não registrado  ....\033[0m")