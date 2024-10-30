import openpyxl
import os
from tabulate import tabulate
import pandas as pd

pd.set_option('future.no_silent_downcasting', True)

#FUNÇÕES PRINCIPAIS

def main():
    while True:

        opcaoMenu = menu()

        if opcaoMenu == 1:
            from vendas import Venda
            Venda.menuVendas()

        elif opcaoMenu == 2 or opcaoMenu == 3:
            from pessoas import Pessoa
            Pessoa.menuPessoas(opcaoMenu)

        elif opcaoMenu == 4:
            from produto import Produto
            Produto.menuProdutos()
            
        else:
            print("Opção invalida....")
            continue

def menu():
    print("\033[1;34m" + "=====================================" + "\033[0m")  
    print("\033[1;32m" + "         Bem-vindo ao Sistema        " + "\033[0m")  
    print("\033[1;34m" + "=====================================" + "\033[0m")  
    print("\033[1;33m" + "[1] Menu Vendas" + "\033[0m")  
    print("\033[1;33m" + "[2] Menu Cliente" + "\033[0m")  
    print("\033[1;33m" + "[3] Menu Vendedor" + "\033[0m")  
    print("\033[1;33m" + "[4] Menu Produtos" + "\033[0m")  
    print("\033[1;34m" + "=====================================" + "\033[0m")

    opcaoMenu = int(input("Selecione uma opção: "))
    
    return opcaoMenu


#FUNÇÕES AUXILIARES

def imprimirTabelas(planilhas, verifPlanilha):
        planilha = pd.read_excel(planilhas, header=None, skiprows= 1)
        cabeçalho = planilha.iloc[0].fillna('').infer_objects(copy=False).tolist()
        planilha.columns = cabeçalho
        planilha = planilha.iloc[1:]
        planilha = planilha.fillna('').infer_objects(copy=False)
        
        tabela = tabulate(planilha, headers='keys', tablefmt='pretty', showindex=False)
        
        if verifPlanilha == 1:
            titulo = "VENDAS"
        elif verifPlanilha == 2:
            titulo = "CLIENTE"
        elif verifPlanilha == 3:
            titulo = "VENDEDOR"
        elif verifPlanilha == 4:
            titulo = "PRODUTOS"
        
        #CALCULAR LARGURA DA PLANILHA PARA CENTRALIZAR O TITULO
        col_widths = [max(planilha[col].astype(str).str.len().max(), len(str(col))) + 2 for col in planilha.columns]
        total_width = sum(col_widths) + len(col_widths) - 1  # -1 para não contar o espaço extra entre colunas

        centeredTitulo = titulo.center(total_width)
        print("\033[1;34m" + centeredTitulo + "\033[0m")
        print(tabela)

def verificarNumeros(verifNumeros):

    if not verifNumeros.isdigit():
            os.system('cls')
            print("Entrada invalida... favor digitar somente números")
            verifNumeros = False
            return verifNumeros

def procurarCelula(planilhas, celulaProcurada):
    celula = None

    for row in range(1, planilhas.max_row + 1):  # Percorre todas as linhas
        for column in range(1, planilhas.max_column + 1):  # Percorre todas as colunas
            if planilhas.cell(row=row, column=column).value == celulaProcurada:
                celula = (row, column)  # Armazena a posição (linha, coluna)
                return celula


if __name__ == '__main__':
    main()














